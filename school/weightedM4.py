import os
import pandas as pd
import numpy as np
import openpyxl
pd.set_option('display.max_rows', None)


def mergeTable():
    # for file in os.listdir('dmt'):
    #     data=pd.read_excel(os.path.join('dmt',file),header=0)
    pass

def insert():
    data = pd.read_excel('dmt/act.xlsx', header=0)
    res=pd.read_excel('dmt/res.xlsx', header=0)
    act=None
    print(data.notnull())
    for row in data.itertuples():
        if isinstance(getattr(row, '活动名称'),str):
            act=getattr(row, '活动名称')
        res.loc[res['姓名'] == row[2],[act]]=3
        # print(res[res['姓名'] == row[2]][act])
    res['学号 ']=res['学号 '].astype(str)
    res.to_excel('dmt/res1.xlsx', header=True,index=False)

def Total():
    res = pd.read_excel('dmt/score.xlsx', header=0)
    res.fillna(value=0,inplace=True)
    print(res.columns)
    res['总计']=0
    for col in ['学生干部', '安全联络员2分', '高分辨率论坛5分', '一带一路青年创新大会5分',
       '高新前沿技术座谈会5分', '科学道德与学风代表3分', '消防演练3分', '扫黑除恶3分', '5月31日校内专家讲座3分',
       '迎新服务活动3分', '毕业班服务活动3分','院庆得分小计','校趣味运动会3分+n*1', '知识竞赛(组织5分，参加3分)',
        '中兴趣味运动会（组织5分，参加3分）', '篮球联赛', '足球联赛','2018.10.10报告会', '2018.10.12报告会', '2018年10.13报告会', '研会成员']:
        res['总计']+=res[col]
    # res['总计']=res['总计'].astype(int)
    print(res['总计'])
    res.to_excel('dmt/score.xlsx',index=False,header=True)

def group():
    g5=['鲍亮', '陈平', '褚华', '杜军朝', '高海昌', '顾新', '黄健斌', '霍秋艳', '李莉', '李青山', '刘惠', '刘西洋', '钱榕', '沈沛意', '史国振', '宋胜利', '覃桂敏', '武波', '谢四江', '李雁妮', '张克君', '张立勇', '张亮', '郑有才', '朱光明', '李瑞', '李贺', '范磊', '董洛兵']

    data=pd.read_excel('dmt/M1M2M4score.xlsx',header=2)
    res = pd.read_excel('dmt/田聪组学生统计.xlsx', header=0)
    print(data.columns)
    index=0
    for col in g5:
        teacher=data[data['导师']==col]
        print(teacher)
        # tem=pd.concat([teacher['学号'],teacher['姓名'],teacher['导师'],teacher['M1模块'],teacher['M2模块'], teacher['M4模块']],axis=1)
        for i in range(len(teacher)):
            row=teacher.iloc[i]
            res.loc[index] =[index+1,row['学号'], row['姓名'], row['导师'], row['M1模块'], row['M2模块'], row['M4模块']]
            index+=1
    res['学号'] = res['学号'].astype(str)
    print(res)
    res.to_excel('dmt/其他组学生统计.xlsx', header=True,index=False)


def colortest():
    filename = "colortest.xlsx"  # 读取excel
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.get_sheet_by_name("试题2")  # 读取Sheet
    rows, cols = worksheet.max_row, worksheet.max_column
    yellow = []
    red = []
    for i in range(1, rows):
        for j in range(1, cols):
            ce = worksheet.cell(row=i, column=j)
            fill = ce.fill
            font = ce.font
            if fill.start_color.rgb == "FFFFFF00" and ce.value != None:
                yellow.append(ce.value)
            if font.color.rgb == "FFFF0000":
                red.append(ce.value)
    print(yellow, red)


def wrost():
    data=pd.read_excel('dmt/2018级M2所有分数.xls',header=0)
    res = pd.read_excel('dmt/软件所学生统计.xlsx', header=0)
    for i in range(len(res)):
        row=res.loc[i]
        stu=data[data['姓名']==row['姓名']]
        import math
        if not math.isnan(stu['不及格门数'].values[0]):
            print(stu['姓名'].values[0],stu['不及格门数'].values[0])
    import numpy as np
    res['学号'] = res['学号'].astype(np.int64).astype(str)
    print((data['学位课平均分']>=75.0))
    ind=~((np.isnan(data['不及格门数']))& (data['学位课平均分']>=75.0))
    data=data[ind]
    data.to_excel('dmt/不及格学生统计.xlsx', header=True,index=False)

    # res.to_excel('dmt/5组学生统计.xlsx', header=True,index=False)

    # print(data['不及格门数'])


def checkStudents():
    data = pd.read_excel('dmt/18级学业奖学金参评名单.xlsx', header=0,converters={'姓名':str,'学号':str},index_col='序号')
    data2=pd.ExcelFile('dmt/2018级导师分组名单最终版.xlsx')
    res=pd.DataFrame(columns=['姓名','学号'])
    for name in data2.sheet_names[1:]:
        _ = pd.read_excel('dmt/2018级导师分组名单最终版.xlsx', sheetname=name, header=0,converters={'姓名':str,'学号':str})
        # print(_.ix[:,['姓名','学号']])
        res=pd.concat([res,_.ix[:,['姓名','学号']]],ignore_index=True)
    res.dropna()
    res['学号']=res['学号'].map(lambda x:str(x).split('.')[0])
    data['学号'] = data['学号'].map(lambda x: x.split('.')[0])

    n=['顾峰', '付楷轩', '葛诗靓', '高琼', '李旭', '刘贺', '陈存粮', '雷昆', '王洪帅', '吕景涛', '强威', '王亚文', '夏祖恒', '黄若航', '刘畅', '李高飞', '曹红兴', '张玉鹏', '陈贵松', '刘家琛', '鱼海洋', '任琳琳', '张建建', '李晨庚', '梁宇', '李成', '李想', '张文祥', '吕晨','梁思宇', '刘洪标', '刘洋', '汪春阳', '王宇晨', '王振翼', '杨健', '姚广宇', '虞小龙', '张鹏', '辛文天', '谈家乐',]

    for index,item in res.iterrows():
       if item['学号'] not in data['学号'].values and item['学号']!='nan':
           if item['姓名'] not in data['姓名'].values and item['姓名']!='Nan' and item['姓名'] not in n:
                print(item['学号'],item['姓名'])

    print('18级学业奖学金参评名单')
    for index,item in data.iterrows():
       if item['学号'] not in res['学号'].values and item['学号']!='nan':
           if item['姓名'] not in res['姓名'].values and item['姓名']!='Nan'and item['姓名'] not in n:
                print(item['学号'],item['姓名'])
    # res=data.merge(res,on=['学号','姓名'],how='outer')
    print(res.__len__())
    print(data.__len__())


if __name__ == '__main__':
    checkStudents()
